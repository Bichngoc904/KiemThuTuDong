import os, csv, json
from openpyxl import load_workbook

def load_data(filename: str, sheet_name: str = None):
    """
    Loader đa năng: đọc dữ liệu từ .md, .csv, .json, .xlsx
    Trả về list[tuple] để dùng cho pytest.mark.parametrize
    :param filename: tên file dữ liệu
    :param sheet_name: tên sheet Excel (nếu có nhiều sheet)
    """
    data = []
    file_path = os.path.join(os.path.dirname(__file__), "..", "data", filename)

    # Markdown
    if filename.endswith(".md"):
        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()
            for line in lines[2:]:  # bỏ header + separator
                row = [x.strip() for x in line.strip().split("|") if x.strip()]
                if row:
                    data.append(tuple(row))

    # CSV
    elif filename.endswith(".csv"):
        with open(file_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)  
            for row in reader:
                data.append(tuple(row))

    # JSON
    elif filename.endswith(".json"):
     with open(file_path, "r", encoding="utf-8") as f:
        json_data = json.load(f)
        for item in json_data:  
            data.append(item) 

    # Excel (.xlsx)
    elif filename.endswith(".xlsx"):
        wb = load_workbook(file_path)
        if sheet_name:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active  # sheet đầu tiên nếu không chỉ định
        if sheet is not None:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(tuple(row))
        else:
            raise ValueError("Workbook does not contain a valid sheet")

    else:
        raise ValueError(f"Chưa hỗ trợ định dạng file này: {filename}")

    return data
